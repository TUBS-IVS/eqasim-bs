"""Extract KBA / MiD fleet reference tables into committed tidy CSVs.

Source files (raw, local-only under ``eqasim-data/data/braunschweig/kba/``,
documented in that directory's ``README.md``):

Legacy xlsx inputs (``kba/`` root, Stichtag 01.01.2025 unless stated):

- ``fz27_202501.xlsx`` (KBA FZ 27 series, stock date 1 January 2025)
    * FZ 27.10 -> segment x powertrain (national)        -> kba_segment_powertrain.csv
    * FZ 27.15 -> Kreis x powertrain (ZGB Kreise only)   -> kba_kreis_powertrain.csv
    * FZ 27.17 -> Gemeinde x private BEV/PHEV (ZGB only)  -> kba_gemeinde_private_bev.csv
    * FZ 27.4  -> Niedersachsen fuel x Euro class         -> kba_fuel_euro_nds.csv
    * FZ 27.7  -> vehicle age band x fuel (Pkw column)    -> kba_age_fuel.csv
    * FZ 27.11 -> brand x powertrain (national)           -> kba_brand_powertrain.csv
- ``fz12_2025.xlsx`` (KBA FZ 12.1; superseded for kba_segment_model.csv)
    * FZ 12.1  -> segment x model (Modellreihe)           (cross-check only)
- ``output_mit_2023_bundesland_fahrzeuge.xlsx`` (MiD 2023)
    * segment x economic status, by Bundesland           -> mid2023_segment_by_status_bundesland.csv
- ``output_mit_2023_raumtyp_fahrzeuge.xlsx`` (MiD 2023)
    * segment x economic status, by RegioStaR Raumtyp     -> mid2023_segment_by_status_raumtyp.csv

Regionalization inputs (``kba/raw/`` subdirectory):

- ``regionalstatistik_46251_02_fuel_kreis_20250101.csv`` (Destatis, Stichtag 01.01.2025)
    * per-Kreis Pkw by fuel type (ALL German Kreise)      -> kba_kreis_fuel.csv
- ``regionalstatistik_46251_03_euro_kreis_20250101.csv`` (Destatis, Stichtag 01.01.2025)
    * per-Kreis Pkw by Euro emission group + Euro-6 substage cols (ALL German Kreise)
                                                          -> kba_kreis_euro.csv
- ``statista_kba_3438_pkw_age_national_2026.xlsx`` (KBA/Statista ID3438, Stichtag 01.01.2026)
    * national Pkw age-band distribution (VALIDATION anchor, not IPF control)
                                                          -> kba_age_national.csv
- ``kba_ev_gemeinde_timeseries_2023_2026.csv`` (KBA, latest period 2026.04)
    * per-Gemeinde EV/BEV/PHEV/fuel-cell share (ZGB only) -> kba_gemeinde_ev.csv
- ``kba_modellreihen_bestand_2020_2026.csv`` (KBA Modellreihen, Stichtag 01.01.2026)
    * per-model-series fuel shares                        -> kba_model_fuel.csv
    * per-segment model share (replacing FZ 12.1)         -> kba_segment_model.csv
- ``kba_ev_grid_5km_2026.gpkg`` (KBA 5 km grid, April 2026)
    * sub-communal EV share tilt (EPSG:3857, ZGB bbox)    -> kba_ev_grid.csv
- ``kba_ev_regiostar7_timeseries_2023_2026.csv`` (KBA, latest period)
    * national EV share by RegioStaR-7 (LOGGING-ONLY cross-check,
      never an IPF control -- see fleet_validation.crosscheck_ev_by_regiostar7)
                                                          -> kba_ev_regiostar7.csv
- ``kba_wohnmobile_holder_age_20250401.csv`` (KBA infographic, COMMITTED)
    * wohnmobile holder-age distribution (hand transcription, issue #315)
                                                          -> kba_wohnmobile_holder_age.csv

The xlsx headers are multi-line and the data starts around row 12, so every
sheet is parsed by *explicit column indices* (documented in the README), never
by header autodetection. KBA / MiD placeholder symbols (``-`` none, ``.``
secret/unknown, ``/`` not reliable, ``()`` uncertain, blank) are coerced to 0
or NaN *explicitly*, and the coercion count per file is logged (project
no-silent-fallback rule).

Every derived CSV that carries provenance data includes a ``stichtag`` column
(ISO date string ``YYYY-MM-DD``) so the reporting date travels with the data.
Multiple Stichtag vintages are present by design (see ADR-0050); do not
reconcile absolute counts across vintages.

The derived CSVs are written under
``eqasim-data/data/braunschweig/kba/derived/`` and force-committed (the
``eqasim-data`` tree is otherwise gitignored), matching the committed-MiD-CSV
pattern. This script is idempotent: re-running it reproduces the same CSVs from
the (unchanged) raw files.

Usage::

    python scripts/extract_kba_fleet.py

Provenance: see ``eqasim-data/data/braunschweig/kba/README.md``. Do not
hand-edit the derived CSVs; re-run this script instead.
"""
from __future__ import annotations

import logging
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

# Running this file directly puts ``scripts/`` on sys.path, not the repository
# root, so the ``braunschweig`` package would be unimportable in the local
# imports further down (extract_gemeinde_ev reuses the production Gemeinde-name
# normaliser). Prepend the repo root explicitly.
_REPO_ROOT = Path(__file__).resolve().parents[1]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

import numpy as np
import openpyxl
import pandas as pd

logging.basicConfig(level=logging.INFO, format="%(message)s")
logger = logging.getLogger("extract_kba_fleet")

# --------------------------------------------------------------------------- #
# Paths
# --------------------------------------------------------------------------- #
KBA_DIR = Path("eqasim-data/data/braunschweig/kba")
DERIVED_DIR = KBA_DIR / "derived"
RAW_DIR = KBA_DIR / "raw"

FZ27_PATH = KBA_DIR / "fz27_202501.xlsx"
FZ12_PATH = KBA_DIR / "fz12_2025.xlsx"
MID_BUNDESLAND_PATH = KBA_DIR / "output_mit_2023_bundesland_fahrzeuge.xlsx"
MID_RAUMTYP_PATH = KBA_DIR / "output_mit_2023_raumtyp_fahrzeuge.xlsx"

FUEL_46251_PATH = RAW_DIR / "regionalstatistik_46251_02_fuel_kreis_20250101.csv"
EURO_46251_PATH = RAW_DIR / "regionalstatistik_46251_03_euro_kreis_20250101.csv"
AGE_NATIONAL_PATH = RAW_DIR / "statista_kba_3438_pkw_age_national_2026.xlsx"
GEMEINDE_EV_PATH = RAW_DIR / "kba_ev_gemeinde_timeseries_2023_2026.csv"
MODELLREIHEN_PATH = RAW_DIR / "kba_modellreihen_bestand_2020_2026.csv"
GRID_EV_PATH = RAW_DIR / "kba_ev_grid_5km_2026.gpkg"
EV_REGIOSTAR7_PATH = RAW_DIR / "kba_ev_regiostar7_timeseries_2023_2026.csv"
WOHNMOBILE_HOLDER_AGE_PATH = RAW_DIR / "kba_wohnmobile_holder_age_20250401.csv"

# --------------------------------------------------------------------------- #
# Canonical label sets
# --------------------------------------------------------------------------- #
# Canonical powertrain labels used everywhere downstream.
POWERTRAIN_LABELS = ("petrol", "diesel", "gas", "bev", "phev", "hybrid", "hydrogen", "other")
# Canonical economic-status labels (MiD 5-class oekonomischer Status).
STATUS_LABELS = ("very_low", "low", "medium", "high", "very_high")
# Canonical snake_case segment labels (stable across KBA + MiD).
SEGMENT_LABELS = (
    "minis", "kleinwagen", "kompaktklasse", "mittelklasse", "obere_mittelklasse",
    "oberklasse", "suv", "gelaendewagen", "sportwagen", "mini_vans",
    "grossraum_vans", "utilities", "wohnmobile", "sonstige",
)

# KBA German segment name -> canonical snake_case segment.
KBA_SEGMENT_MAP = {
    "minis": "minis",
    "kleinwagen": "kleinwagen",
    "kompaktklasse": "kompaktklasse",
    "mittelklasse": "mittelklasse",
    "obere mittelklasse": "obere_mittelklasse",
    "oberklasse": "oberklasse",
    "suvs": "suv",
    "gelaendewagen": "gelaendewagen",
    "gelandewagen": "gelaendewagen",
    "sportwagen": "sportwagen",
    "mini-vans": "mini_vans",
    "grossraum-vans": "grossraum_vans",
    "grosraum-vans": "grossraum_vans",
    "utilities": "utilities",
    "wohnmobile": "wohnmobile",
    "sonstige": "sonstige",
}

def _report_unmapped_segments(where: str, unmapped_stock: "dict[str, float]",
                              mapped_stock: float) -> None:
    """Log the segment labels that did not map, with their stock share.

    Naming the offending LABELS is the actionable part: a bare skipped-row count
    reads like a tolerable residual, while the label plus its stock share says
    whether a whole segment just vanished. Skipping an unknown label is legitimate
    on its own (a source may carry an "Unbekannt" residual); what is NOT
    legitimate is a canonical segment ending up without model rows, and that is
    checked against the segment marginal in :func:`_check_segment_model_coverage`.
    """
    if not unmapped_stock:
        return
    total = mapped_stock + sum(unmapped_stock.values())
    share = sum(unmapped_stock.values()) / total if total > 0 else 0.0
    detail = ", ".join(
        f"{label!r} ({stock:,.0f}, {100.0 * stock / total:.2f}%)"
        for label, stock in sorted(unmapped_stock.items(),
                                   key=lambda kv: -kv[1])
    )
    logger.warning(
        "[%s] %d segment label(s) did not map via KBA_SEGMENT_MAP and were "
        "skipped: %s; that is %.2f%% of the source stock. Add a label to "
        "KBA_SEGMENT_MAP (or extend segment_lookup_key if it is only a spelling "
        "variant) if one of these should have been kept.",
        where, len(unmapped_stock), detail, 100.0 * share,
    )


def _check_segment_model_coverage(df_segment_model: pd.DataFrame,
                                  df_segment_powertrain: pd.DataFrame) -> None:
    """Every segment of the segment MARGINAL must have model rows.

    The segment marginal (FZ 27.10 -> ``kba_segment_powertrain.csv``) decides how
    many cars are drawn into each segment; the model table
    (``kba_segment_model.csv``) supplies the brand/model pool for that segment. A
    segment present in the first but missing from the second is not a tolerable
    data gap: every car drawn into it gets an EMPTY brand and model, and with it no
    HSN/TSN engine attributes and no per-model fuel weighting.

    Measured trigger (issue #277): the 2026 Modellreihen export writes
    ``"GELÄNDEWAGEN"`` with an umlaut. The ASCII-only key lookup missed it, so all
    56 Geländewagen model rows were skipped while FZ 27.10 kept the segment --
    5.6% of drawn cars ended up with an empty brand, silently.
    """
    marginal = set(df_segment_powertrain["segment"].astype(str))
    modelled = set(df_segment_model["segment"].astype(str))
    missing = sorted(marginal - modelled)
    if missing:
        raise ValueError(
            f"segments present in the segment marginal but absent from "
            f"kba_segment_model.csv: {missing}. Every car drawn into them would "
            "carry an empty brand/model. Check KBA_SEGMENT_MAP / "
            "segment_lookup_key against the raw source's segment spellings."
        )
    logger.info(
        "[segment coverage] all %d marginal segment(s) have model rows "
        "(model table covers %d segment(s))", len(marginal), len(modelled),
    )


def segment_lookup_key(label: str) -> str:
    """Normalise a source segment label for the segment-map lookup.

    Lower-cases, folds German umlauts/eszett to the ASCII spelling the maps use
    and collapses whitespace, so a source that writes ``"GELÄNDEWAGEN"``,
    ``"Gelaendewagen"`` or ``"gelandewagen"`` all reach the same entry. Applied at
    EVERY segment-map lookup; a map miss then means a genuinely unknown segment,
    not a spelling variant.
    """
    text = str(label).strip().lower()
    for umlaut, ascii_form in (("ä", "ae"), ("ö", "oe"), ("ü", "ue"), ("ß", "ss")):
        text = text.replace(umlaut, ascii_form)
    return " ".join(text.split())


# MiD German segment name -> canonical snake_case segment.
# MiD uses "Sportgelaendewagen" for what KBA calls SUVs, and
# "nicht zuzuordnen" for the residual ("Sonstige").  MiD has no Wohnmobile block.
MID_SEGMENT_MAP = dict(KBA_SEGMENT_MAP)
MID_SEGMENT_MAP.update({
    "sportgelaendewagen": "suv",
    "nicht zuzuordnen": "sonstige",
})

# MiD region header labels arrive line-wrapped (e.g. "Niedersach sen",
# "Baden- Wuerttember g", "kleinstaedtische r, doerflicher Raum").  Map the
# wrapped (whitespace-collapsed, umlaut-normalised) form to a clean label.
MID_BUNDESLAND_NAME_MAP = {
    "Schleswig- Holstein": "Schleswig-Holstein",
    "Hamburg": "Hamburg",
    "Niedersach sen": "Niedersachsen",
    "Bremen": "Bremen",
    "Nordrhein- Westfalen": "Nordrhein-Westfalen",
    "Hessen": "Hessen",
    "Rheinland- Pfalz": "Rheinland-Pfalz",
    "Baden- Wuerttember g": "Baden-Wuerttemberg",
    "Bayern": "Bayern",
    "Saarland": "Saarland",
    "Berlin": "Berlin",
    "Brandenbu rg": "Brandenburg",
    "Mecklenbur g- Vorpommer n": "Mecklenburg-Vorpommern",
    "Sachsen": "Sachsen",
    "Sachsen- Anhalt": "Sachsen-Anhalt",
    "Thueringen": "Thueringen",
}
MID_RAUMTYP_NAME_MAP = {
    "Stadtregion - Metropole": "Stadtregion - Metropole",
    "Stadtregion - Regiopole und Grossstadt": "Stadtregion - Regiopole und Grossstadt",
    "Stadtregion - Mittelstadt, staedtischer Raum": "Stadtregion - Mittelstadt, staedtischer Raum",
    "Stadtregion - kleinstaedtische r, doerflicher Raum": "Stadtregion - kleinstaedtischer, doerflicher Raum",
    "laendliche Region - zentrale Stadt": "laendliche Region - zentrale Stadt",
    "laendliche Region - Mittelstadt, staedtischer Raum": "laendliche Region - Mittelstadt, staedtischer Raum",
    "laendliche Region - kleinstaedtische r, doerflicher Raum": "laendliche Region - kleinstaedtischer, doerflicher Raum",
}

# MiD German economic-status label -> canonical status.
MID_STATUS_MAP = {
    "sehr niedrig": "very_low",
    "niedrig": "low",
    "mittel": "medium",
    "hoch": "high",
    "sehr hoch": "very_high",
}

# Wohnmobile holder-age classes (issue #315). Deliberately duplicated in
# braunschweig.data.kba.fleet_tables (mutual extractor<->loader contract).
WOHNMOBILE_AGE_CLASS_LABELS = (
    "up_to_20", "21_29", "30_39", "40_49", "50_59", "60_69", "70_79", "80_plus",
)
WOHNMOBILE_AGE_NOT_ATTRIBUTED = "not_attributed"

# KBA fuel ("Kraftstoffart") label -> canonical powertrain (FZ 27.4 / FZ 27.7).
KBA_FUEL_MAP = {
    "benzin": "petrol",
    "diesel": "diesel",
    "gas insgesamt": "gas",
    "elektro (bev)": "bev",
    "hybrid insgesamt": "hybrid",
    "darunter plug-in": "phev",
    "sonstige": "other",
}

# The 7 RegioStaR-7 codes (BMV/BBSR typology; 99 = "keine Zuordnung" is dropped).
RS7_CODES = (71, 72, 73, 74, 75, 76, 77)

# ZGB Kreise (KBA Kennziffer == Kreis AGS-5 == "03" + Kreis3).
ZGB_KREISE = {
    "03101": "Braunschweig, Stadt",
    "03102": "Salzgitter",
    "03103": "Wolfsburg",
    "03151": "Gifhorn",
    "03153": "Goslar",
    "03154": "Helmstedt",
    "03157": "Peine",
    "03158": "Wolfenbuettel",
}

# German Statista age-band labels -> canonical snake_case band labels (ID 3438).
# These 6 bands cover the full Pkw fleet; they are a VALIDATION control, never
# an IPF dimension.
_AGE_NATIONAL_BANDS = {
    "unter 2 jahre": "under_2",
    "2 bis 4 jahre": "2_to_4",
    "5 bis 9 jahre": "5_to_9",
    "10 bis 14 jahre": "10_to_14",
    "15 bis 29 jahre": "15_to_29",
    "30 und mehr jahre": "30_plus",
}

# KBA placeholder symbols that map to a numeric 0 (count below 0.5 / none).
_ZERO_SYMBOLS = {"-", "0"}
# KBA placeholder symbols that map to NaN (secret / not reliable / uncertain).
_NAN_SYMBOLS = {".", "/", "()", "(", ")", "x", "X"}


@dataclass
class CoercionCounter:
    """Tracks how many placeholder cells were coerced to 0 / NaN per file."""

    file_label: str
    to_zero: int = 0
    to_nan: int = 0
    cells_seen: int = 0
    distinct_symbols: dict = field(default_factory=dict)

    def record(self, raw_value, coerced_value) -> None:
        self.cells_seen += 1
        if isinstance(raw_value, str):
            symbol = raw_value.strip()
        else:
            symbol = raw_value
        if pd.isna(coerced_value):
            self.to_nan += 1
            self.distinct_symbols[str(symbol)] = self.distinct_symbols.get(str(symbol), 0) + 1
        elif coerced_value == 0 and not _looks_numeric_zero(raw_value):
            self.to_zero += 1
            self.distinct_symbols[str(symbol)] = self.distinct_symbols.get(str(symbol), 0) + 1

    def log(self) -> None:
        total = self.to_zero + self.to_nan
        logger.info(
            "[%s] numeric cells=%d, coerced placeholders=%d (-> 0: %d, -> NaN: %d); symbols=%s",
            self.file_label, self.cells_seen, total, self.to_zero, self.to_nan,
            dict(sorted(self.distinct_symbols.items())),
        )


def _looks_numeric_zero(raw_value) -> bool:
    """True if the raw value is a genuine numeric 0, not a placeholder string."""
    return isinstance(raw_value, (int, float)) and not isinstance(raw_value, bool) and raw_value == 0


def _coerce_count(raw_value, counter: CoercionCounter):
    """Coerce a KBA/MiD count cell to float, mapping placeholder symbols explicitly.

    ``- 0`` (string) -> 0; ``. / ()`` / blank -> NaN; numbers pass through.
    Every coercion of a placeholder symbol is recorded in ``counter``.
    """
    if raw_value is None:
        result = np.nan
        counter.record(raw_value, result)
        return result
    if isinstance(raw_value, (int, float)) and not isinstance(raw_value, bool):
        # Genuine numeric value (including a real 0); not a placeholder.
        counter.cells_seen += 1
        return float(raw_value)
    text = str(raw_value).strip()
    if text == "":
        result = np.nan
        counter.record(raw_value, result)
        return result
    if text in _ZERO_SYMBOLS:
        result = 0.0
        counter.record(raw_value, result)
        return result
    if text in _NAN_SYMBOLS:
        result = np.nan
        counter.record(raw_value, result)
        return result
    # German thousands separators (".") inside numbers, decimal comma.
    cleaned = text.replace(".", "").replace(",", ".")
    try:
        value = float(cleaned)
        counter.cells_seen += 1
        return value
    except ValueError:
        # Unknown non-numeric token -> NaN, recorded as a coercion.
        result = np.nan
        counter.record(raw_value, result)
        return result


def _coerce_percent(raw_value, counter: CoercionCounter):
    """Coerce a MiD percent cell like ``'7 %'`` to a float share in percent.

    ``'7 %'`` -> 7.0; placeholders -> NaN (recorded).
    """
    if raw_value is None:
        result = np.nan
        counter.record(raw_value, result)
        return result
    if isinstance(raw_value, (int, float)) and not isinstance(raw_value, bool):
        counter.cells_seen += 1
        return float(raw_value)
    text = str(raw_value).strip()
    if text == "" or text in _NAN_SYMBOLS:
        result = np.nan
        counter.record(raw_value, result)
        return result
    cleaned = text.replace("%", "").replace(",", ".").strip()
    try:
        value = float(cleaned)
        counter.cells_seen += 1
        return value
    except ValueError:
        result = np.nan
        counter.record(raw_value, result)
        return result


def _normalise_label(text: str) -> str:
    """Collapse line-wrap whitespace and German umlauts for label matching."""
    if text is None:
        return ""
    collapsed = " ".join(str(text).split())
    return (
        collapsed.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue")
        .replace("Ä", "Ae").replace("Ö", "Oe").replace("Ü", "Ue")
        .replace("ß", "ss")
    )


def _cell(row, index):
    """Return ``row[index]`` or None when the row is shorter (ragged xlsx rows)."""
    return row[index] if index < len(row) else None


def _read_sheet(path: Path, sheet_name: str):
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    workbook.close()
    return rows


# --------------------------------------------------------------------------- #
# FZ 27.10 -> kba_segment_powertrain.csv
# --------------------------------------------------------------------------- #
def extract_segment_powertrain() -> pd.DataFrame:
    """FZ 27.10: per-segment total + BEV/PHEV/hybrid/gas/hydrogen counts.

    Column indices (0-based, verified against the data rows): col1=segment,
    col2=total, col3=alt total, col5=electro total, col7=BEV (Elektro),
    col8=fuel cell (hydrogen, in the electro group), col9=PHEV, col10=hybrid
    total, col13=gas insgesamt, col14=hydrogen (Wasserstoff column).  Only the
    14 KBA segment rows of the 1 January 2025 block are read (the first block;
    later blocks repeat prior years).
    """
    rows = _read_sheet(FZ27_PATH, "FZ 27.10")
    counter = CoercionCounter("FZ 27.10 segment_powertrain")
    records = []
    started = False
    for row in rows:
        label = _normalise_label(row[1]).lower() if row[1] is not None else ""
        if not started:
            # The 1 January 2025 total row precedes the segment rows.
            if label.startswith("1. januar 2025"):
                started = True
            continue
        if label.startswith("1. januar"):
            break  # next year's block
        canonical = KBA_SEGMENT_MAP.get(segment_lookup_key(label))
        if canonical is None:
            continue
        total = _coerce_count(_cell(row, 2), counter)
        # F9: a KBA-suppressed alt-drive cell (placeholder "." / "()" / blank)
        # coerces to NaN; wrap every alt-drive count in np.nan_to_num so a
        # suppressed cell cannot propagate NaN into alt_total / *_share below
        # (the hydrogen column is already NaN-safe via np.nansum -- apply the
        # same guard to bev/phev/hybrid/gas for consistency).
        bev = np.nan_to_num(_coerce_count(_cell(row, 7), counter))
        fuel_cell = _coerce_count(_cell(row, 8), counter)
        phev = np.nan_to_num(_coerce_count(_cell(row, 9), counter))
        hybrid = np.nan_to_num(_coerce_count(_cell(row, 10), counter))
        gas = np.nan_to_num(_coerce_count(_cell(row, 13), counter))
        hydrogen_col = _coerce_count(_cell(row, 14), counter)
        hydrogen = np.nansum([fuel_cell, hydrogen_col])
        records.append({
            "segment": canonical,
            "total": total,
            "bev": bev,
            "phev": phev,
            "hybrid": hybrid,
            "gas": gas,
            "hydrogen": hydrogen,
        })
    counter.log()
    frame = pd.DataFrame(records)
    # The "alternative drive" columns are a strict subset of the total; the
    # remaining vehicles are conventional petrol/diesel (not separable in this
    # sheet).  We expose the alternative counts plus their per-segment shares.
    alt_total = frame[["bev", "phev", "hybrid", "gas", "hydrogen"]].sum(axis=1)
    frame["alt_total"] = alt_total
    frame["segment_share"] = frame["total"] / frame["total"].sum()
    for powertrain in ("bev", "phev", "hybrid", "gas", "hydrogen"):
        frame[f"{powertrain}_share"] = frame[powertrain] / frame["total"]
    return frame


# --------------------------------------------------------------------------- #
# FZ 27.15 -> kba_kreis_powertrain.csv (ZGB Kreise only)
# --------------------------------------------------------------------------- #
def extract_kreis_powertrain() -> pd.DataFrame:
    """FZ 27.15: per-Kreis total + alternative-drive + BEV/PHEV/hybrid/gas.

    Column indices (0-based, verified against the data rows): col2=Kennziffer
    (= Kreis AGS-5), col3=name, col4=total, col5=alt total, col7=electro total,
    col9=BEV, col10=PHEV, col11=hybrid total, col14=gas insgesamt (the data
    rows have no hydrogen column, so gas is the last cell).  Filtered to the 8
    ZGB Kreise.
    """
    rows = _read_sheet(FZ27_PATH, "FZ 27.15")
    counter = CoercionCounter("FZ 27.15 kreis_powertrain")
    records = []
    for row in rows:
        kennziffer = _cell(row, 2)
        if kennziffer is None:
            continue
        code = str(kennziffer).strip()
        if code not in ZGB_KREISE:
            continue
        # F9: guard every alt-drive count against a KBA-suppressed cell
        # (placeholder -> NaN via _coerce_count) so it cannot propagate NaN
        # into alt_share/bev_share/phev_share below.
        records.append({
            "kreis_ags5": code,
            "kreis_name": ZGB_KREISE[code],
            "total": _coerce_count(_cell(row, 4), counter),
            "alt_total": np.nan_to_num(_coerce_count(_cell(row, 5), counter)),
            "bev": np.nan_to_num(_coerce_count(_cell(row, 9), counter)),
            "phev": np.nan_to_num(_coerce_count(_cell(row, 10), counter)),
            "hybrid": np.nan_to_num(_coerce_count(_cell(row, 11), counter)),
            "gas": np.nan_to_num(_coerce_count(_cell(row, 14), counter)),
        })
    counter.log()
    frame = pd.DataFrame(records).sort_values("kreis_ags5").reset_index(drop=True)
    frame["bev_share"] = frame["bev"] / frame["total"]
    frame["phev_share"] = frame["phev"] / frame["total"]
    frame["alt_share"] = frame["alt_total"] / frame["total"]
    return frame


# --------------------------------------------------------------------------- #
# FZ 27.17 -> kba_gemeinde_private_bev.csv (ZGB Kreise only)
# --------------------------------------------------------------------------- #
def extract_gemeinde_private_bev() -> pd.DataFrame:
    """FZ 27.17: per-Gemeinde private car total + private BEV/PHEV.

    Column indices (0-based, README): col2=Zulassungsbezirk ("AGS5 NAME",
    filled once per Bezirk), col3=Gemeinde, col7=private total, col8=private
    BEV, col9=private PHEV.  Filtered to Gemeinden whose Bezirk prefix is a ZGB
    Kreis AGS-5.
    """
    rows = _read_sheet(FZ27_PATH, "FZ 27.17")
    counter = CoercionCounter("FZ 27.17 gemeinde_private_bev")
    records = []
    current_kreis: Optional[str] = None
    for row in rows:
        bezirk = _cell(row, 2)
        if bezirk is not None and str(bezirk).strip():
            # e.g. "08111 STUTTGART,STADT" -> the leading 5-digit AGS is the Kreis.
            token = str(bezirk).strip().split()[0]
            if token.isdigit() and len(token) == 5:
                current_kreis = token
        gemeinde = _cell(row, 3)
        if gemeinde is None or not str(gemeinde).strip():
            continue
        if current_kreis not in ZGB_KREISE:
            continue
        records.append({
            "kreis_ags5": current_kreis,
            "kreis_name": ZGB_KREISE[current_kreis],
            "gemeinde": str(gemeinde).strip(),
            "private_total": _coerce_count(_cell(row, 7), counter),
            "private_bev": _coerce_count(_cell(row, 8), counter),
            "private_phev": _coerce_count(_cell(row, 9), counter),
        })
    counter.log()
    frame = pd.DataFrame(records)
    frame = frame.sort_values(["kreis_ags5", "gemeinde"]).reset_index(drop=True)
    frame["private_bev_share"] = frame["private_bev"] / frame["private_total"]
    frame["private_phev_share"] = frame["private_phev"] / frame["private_total"]
    return frame


# --------------------------------------------------------------------------- #
# FZ 27.4 -> kba_fuel_euro_nds.csv (Niedersachsen)
# --------------------------------------------------------------------------- #
def extract_fuel_euro_nds() -> pd.DataFrame:
    """FZ 27.4: Niedersachsen fuel x Euro class (long form).

    Column indices (0-based, README): col1=Land (once per block), col2=fuel,
    col3=Euro1, col4=Euro2, col5=Euro3, col6=Euro4, col7=Euro5,
    col8=Euro6 total, col12=Sonstige, col13=row total.  We keep the headline
    Euro classes (1..6, plus "other" = Sonstige) and skip the Euro-6
    sub-breakdown columns (6d-temp / 6d / 6e) so the per-fuel Euro shares sum
    to one without double counting.
    """
    rows = _read_sheet(FZ27_PATH, "FZ 27.4")
    counter = CoercionCounter("FZ 27.4 fuel_euro_nds")
    euro_columns = {
        "euro1": 3, "euro2": 4, "euro3": 5, "euro4": 6, "euro5": 7,
        "euro6": 8, "other": 12,
    }
    records = []
    in_block = False
    for row in rows:
        land = _normalise_label(row[1]) if row[1] is not None else ""
        if land.lower().startswith("niedersachsen zusammen"):
            break
        if land.lower() == "niedersachsen":
            in_block = True
        if not in_block:
            continue
        fuel_label = _normalise_label(row[2]).lower() if row[2] is not None else ""
        powertrain = KBA_FUEL_MAP.get(fuel_label)
        if powertrain is None:
            continue
        for euro_class, col in euro_columns.items():
            records.append({
                "fuel": powertrain,
                "euro_class": euro_class,
                "count": _coerce_count(_cell(row, col), counter),
            })
    counter.log()
    frame = pd.DataFrame(records)
    # Per-fuel Euro-class share (within powertrain).
    fuel_totals = frame.groupby("fuel")["count"].transform("sum")
    frame["share"] = frame["count"] / fuel_totals
    return frame


# --------------------------------------------------------------------------- #
# FZ 27.4 -> kba_fuel_euro6_substage_nds.csv (Niedersachsen, Task B4)
# --------------------------------------------------------------------------- #
def extract_fuel_euro6_substage_nds() -> pd.DataFrame:
    """FZ 27.4: Niedersachsen fuel x Euro-6 substage breakdown (long form).

    :func:`extract_fuel_euro_nds` keeps only the headline Euro classes
    (1..6, other) and skips the Euro-6 sub-breakdown columns. This function
    extracts that sub-breakdown instead, using the SAME column layout (see
    that function's docstring): col1=Land (once per block), col2=fuel,
    col8=Euro6 total, col9=darunter Euro-6d-temp, col10=darunter Euro-6d,
    col11=darunter Euro-6e.

    Euro-6e is the newest sub-class and has no separate downstream bucket (the
    HBEFA Euro-6 sub-mapping in Task B5 only distinguishes 6a/b/c, 6d-temp and
    6d), so it is FOLDED INTO Euro-6d here -- a NaN-safe fold so a KBA-
    suppressed or absent 6e cell does not turn the residual computation
    negative. The residual ("Euro-6a/b/c", not reported directly by Destatis)
    is derived as ``euro6ab = max(euro6_total - euro6dtemp - euro6d, 0)`` per
    fuel, mirroring the per-Kreis 46251-03 derivation in
    :func:`extract_kreis_euro_46251`.

    Returns:
        Long-form DataFrame with columns ``fuel, substage, count, share,
        stichtag``, where ``substage`` in ``{"euro6ab", "euro6dtemp",
        "euro6d"}`` and ``share = P(substage | euro6, fuel)`` (the three
        substage counts sum to the fuel's euro6 total, so shares sum to 1
        within each fuel that has a positive euro6 total).
    """
    rows = _read_sheet(FZ27_PATH, "FZ 27.4")
    counter = CoercionCounter("FZ 27.4 fuel_euro6_substage_nds")
    records = []
    in_block = False
    for row in rows:
        land = _normalise_label(row[1]) if row[1] is not None else ""
        if land.lower().startswith("niedersachsen zusammen"):
            break
        if land.lower() == "niedersachsen":
            in_block = True
        if not in_block:
            continue
        fuel_label = _normalise_label(row[2]).lower() if row[2] is not None else ""
        powertrain = KBA_FUEL_MAP.get(fuel_label)
        if powertrain is None:
            continue
        euro6_total = _coerce_count(_cell(row, 8), counter)
        euro6dtemp = np.nan_to_num(_coerce_count(_cell(row, 9), counter))
        euro6d_raw = np.nan_to_num(_coerce_count(_cell(row, 10), counter))
        euro6e = np.nan_to_num(_coerce_count(_cell(row, 11), counter))
        euro6d = euro6d_raw + euro6e  # fold Euro-6e into Euro-6d (documented above)
        euro6ab = max(np.nan_to_num(euro6_total) - euro6dtemp - euro6d, 0.0)
        records.append({"fuel": powertrain, "substage": "euro6ab", "count": euro6ab})
        records.append({"fuel": powertrain, "substage": "euro6dtemp", "count": euro6dtemp})
        records.append({"fuel": powertrain, "substage": "euro6d", "count": euro6d})
    counter.log()
    frame = pd.DataFrame(records)
    fuel_totals = frame.groupby("fuel")["count"].transform("sum")
    # A fuel with ZERO Euro-6 registrations (plausible for a low-volume fuel such
    # as gas/phev in a Land-level cut) has a zero total; ``count / 0`` would be a
    # silent NaN. Divide by a floored denominator so those rows get share 0.0
    # instead of NaN (no-NA / no-silent-fallback rule), and log how many rows are
    # affected. Their counts are all 0, so 0/1.0 == 0.0 is exact. The B5 substage
    # draw treats an all-zero fuel substage pmf as a fallback (national euro6ab).
    safe_totals = fuel_totals.where(fuel_totals > 0, other=1.0)
    frame["share"] = frame["count"] / safe_totals
    n_zero_total = int((fuel_totals == 0).sum())
    if n_zero_total:
        logger.warning(
            "[extract_fuel_euro6_substage_nds] %d substage rows have a zero euro6 "
            "fuel total (fuel with no Euro-6 registrations) -> share set to 0.0.",
            n_zero_total,
        )
    frame["stichtag"] = "2025-01-01"
    return frame


# --------------------------------------------------------------------------- #
# FZ 27.7 -> kba_age_fuel.csv (Pkw column)
# --------------------------------------------------------------------------- #
def extract_age_fuel() -> pd.DataFrame:
    """FZ 27.7: vehicle age band x fuel, Pkw (passenger-car) column only.

    Column indices (0-based, README): col1=Fahrzeugalter (filled once per
    block), col2=Kraftstoffart, col4=Personenkraftwagen (the Pkw column).
    "... zusammen" subtotal rows are skipped.
    """
    rows = _read_sheet(FZ27_PATH, "FZ 27.7")
    counter = CoercionCounter("FZ 27.7 age_fuel")
    records = []
    current_age: Optional[str] = None
    for row in rows:
        age_label = _normalise_label(row[1]) if row[1] is not None else ""
        if age_label and "zusammen" in age_label.lower():
            current_age = None
            continue
        if age_label and "Jahre" in age_label:
            current_age = _canonical_age_band(age_label)
        fuel_label = _normalise_label(row[2]).lower() if row[2] is not None else ""
        powertrain = KBA_FUEL_MAP.get(fuel_label)
        if powertrain is None or current_age is None:
            continue
        records.append({
            "age_band": current_age,
            "fuel": powertrain,
            "pkw_count": _coerce_count(_cell(row, 4), counter),
        })
    counter.log()
    frame = pd.DataFrame(records)
    fuel_totals = frame.groupby("fuel")["pkw_count"].transform("sum")
    frame["share"] = frame["pkw_count"] / fuel_totals
    return frame


def _canonical_age_band(german_label: str) -> str:
    """Map a German FZ 27.7 age band to a stable snake_case label."""
    text = german_label.lower()
    mapping = {
        "unter 5 jahre": "under_5",
        "5 bis 9 jahre": "5_to_9",
        "10 bis 14 jahre": "10_to_14",
        "15 bis 19 jahre": "15_to_19",
        "20 bis 24 jahre": "20_to_24",
        "25 bis 29 jahre": "25_to_29",
        "30 jahre und mehr": "30_plus",
    }
    return mapping.get(text, text.replace(" ", "_"))


# --------------------------------------------------------------------------- #
# FZ 27.11 -> kba_brand_powertrain.csv
# --------------------------------------------------------------------------- #
def extract_brand_powertrain() -> pd.DataFrame:
    """FZ 27.11: per-brand total + BEV/PHEV/hybrid/gas counts.

    Same column layout as FZ 27.10: col1=brand, col2=total, col7=BEV,
    col9=PHEV, col10=hybrid total, col13=gas insgesamt.  The trailing
    ``INSGESAMT`` grand-total row is skipped; ``SONSTIGE`` (residual brands) is
    kept.
    """
    rows = _read_sheet(FZ27_PATH, "FZ 27.11")
    counter = CoercionCounter("FZ 27.11 brand_powertrain")
    records = []
    started = False
    for row in rows:
        label = _normalise_label(row[1]) if row[1] is not None else ""
        if not label:
            continue
        if not started:
            # Brand rows begin after the column-header block; the first brand
            # row has a numeric total in col2.
            if isinstance(_cell(row, 2), (int, float)) and label.upper() != "INSGESAMT":
                started = True
            else:
                continue
        if label.upper() == "INSGESAMT":
            break
        if "Kraftfahrt-Bundesamt" in label:
            break
        records.append({
            "brand": label,
            "total": _coerce_count(_cell(row, 2), counter),
            "bev": _coerce_count(_cell(row, 7), counter),
            "phev": _coerce_count(_cell(row, 9), counter),
            "hybrid": _coerce_count(_cell(row, 10), counter),
            "gas": _coerce_count(_cell(row, 13), counter),
        })
    counter.log()
    frame = pd.DataFrame(records)
    frame["brand_share"] = frame["total"] / frame["total"].sum()
    return frame


# --------------------------------------------------------------------------- #
# FZ 12.1 -> kba_segment_model.csv
# --------------------------------------------------------------------------- #
def extract_segment_model() -> pd.DataFrame:
    """FZ 12.1: per-segment model (Modellreihe) counts and within-segment share.

    .. deprecated::
        Superseded by :func:`extract_segment_model_2026` which reads the newer
        Modellreihen bestand CSV (01.01.2026); kept for reference/reversibility.

    Column indices (0-based, README): col1=Segment (filled once per block,
    uppercase), col2=Modellreihe, col3=Anzahl (count).  Segment subtotal rows
    ("... ZUSAMMEN") and the trailing ``BESTAND INSGESAMT`` / footnotes are
    skipped.  The within-segment share is recomputed from the counts so it is
    exact and reproducible.
    """
    rows = _read_sheet(FZ12_PATH, "FZ 12.1")
    counter = CoercionCounter("FZ 12.1 segment_model")
    records = []
    current_segment: Optional[str] = None
    for row in rows:
        seg_label = _normalise_label(row[1]) if row[1] is not None else ""
        if seg_label:
            upper = seg_label.upper().strip()
            if "ZUSAMMEN" in upper:
                current_segment = None
                continue
            # End of the data block (grand total + footnotes).  Match only the
            # exact "BESTAND INSGESAMT" / "HINWEIS:" markers, never the sheet
            # title row ("Bestand an Personenkraftwagen nach Segmenten ...").
            if upper.startswith("BESTAND INSGESAMT") or upper.startswith("HINWEIS"):
                break
            if "Kraftfahrt-Bundesamt" in seg_label:
                break
            canonical = KBA_SEGMENT_MAP.get(segment_lookup_key(seg_label))
            if canonical is not None:
                current_segment = canonical
        model = _cell(row, 2)
        if model is None or not str(model).strip():
            continue
        if current_segment is None:
            continue
        records.append({
            "segment": current_segment,
            "model": str(model).strip(),
            "count": _coerce_count(_cell(row, 3), counter),
        })
    counter.log()
    frame = pd.DataFrame(records)
    segment_totals = frame.groupby("segment")["count"].transform("sum")
    frame["share"] = frame["count"] / segment_totals
    return frame


# --------------------------------------------------------------------------- #
# kba_modellreihen_bestand_2020_2026.csv helpers
# --------------------------------------------------------------------------- #
#: Columns the Modellreihen export must carry for the downstream extractors.
MODELLREIHEN_REQUIRED_COLUMNS = ("Berichtszeitpunkt", "Segment", "Marke",
                                 "Modellreihe", "Anzahl")


def _read_modellreihen(path) -> pd.DataFrame:
    """Read the KBA Modellreihen CSV (utf-8-sig), filter to 01.01.2026.

    Returns the raw filtered DataFrame.  The CSV must carry columns:
    ``Berichtszeitpunkt, Segment, Marke, Modellreihe, Anzahl, Diesel,
    Hybrid, Hybrid_Plugin, BEV, gewerblich``.

    The delimiter is SNIFFED rather than assumed: the KBA portal has shipped this
    export both comma-separated (the 2020-2026 file in ``kba/raw``) and
    semicolon-separated. A hard-coded separator loads the whole file into a
    single column and fails later with an opaque ``KeyError`` on the first
    expected column name, so the resolved column set is validated here and the
    error names the file and what was actually parsed.

    Args:
        path: Path-like or str pointing at the raw CSV.

    Returns:
        DataFrame with only the 01.01.2026 rows.

    Raises:
        ValueError: if the parsed frame lacks any of
            :data:`MODELLREIHEN_REQUIRED_COLUMNS`.
    """
    # sep=None + the python engine runs csv.Sniffer over the header.
    raw = pd.read_csv(path, sep=None, engine="python", encoding="utf-8-sig",
                      dtype=str)
    # Normalise column names: strip whitespace AND any byte-order mark. The BOM
    # strip is not redundant with encoding="utf-8-sig": the pandas C parser drops
    # a BOM that survives decoding, the python engine (required for delimiter
    # sniffing) does not, and a file whose content itself starts with U+FEFF and
    # is additionally written as utf-8-sig carries two -- decoding removes one and
    # the other would end up inside the first column's name.
    raw.columns = [c.strip().lstrip("﻿").strip() for c in raw.columns]
    missing = [c for c in MODELLREIHEN_REQUIRED_COLUMNS if c not in raw.columns]
    if missing:
        raise ValueError(
            f"{path}: Modellreihen export is missing the required column(s) "
            f"{missing}. Parsed {len(raw.columns)} column(s): "
            f"{list(raw.columns)[:8]}{' ...' if len(raw.columns) > 8 else ''}. "
            "A single parsed column means the delimiter could not be sniffed -- "
            "check that the file is a KBA Modellreihen Bestand export."
        )
    logger.info("[modellreihen] %s: parsed %d columns, %d rows",
                path, len(raw.columns), len(raw))
    filtered = raw[raw["Berichtszeitpunkt"].str.strip() == "01.01.2026"].copy()
    logger.info("[modellreihen] %d rows at Stichtag 01.01.2026", len(filtered))
    return filtered


# --------------------------------------------------------------------------- #
# Modellreihen -> kba_segment_model.csv (2026 refresh)
# --------------------------------------------------------------------------- #
def extract_segment_model_2026(path=None) -> pd.DataFrame:
    """Modellreihen (2026): per-segment model counts + within-segment share.

    Reads the KBA Modellreihen bestand CSV filtered to Berichtszeitpunkt
    ``01.01.2026`` and produces the same schema as the legacy
    ``extract_segment_model`` (``segment, model, count, share``) plus a
    ``stichtag`` column.

    The join key convention matches ``kba_segment_model.csv`` from FZ 12.1:
    ``model = f"{Marke} {Modellreihe}"`` (uppercase as delivered by KBA).

    Rows whose ``Segment`` does not map via ``KBA_SEGMENT_MAP`` are skipped;
    the skip count is logged (no-silent-fallback rule).

    Args:
        path: Path to the raw Modellreihen CSV.  Defaults to
            ``MODELLREIHEN_PATH`` when ``None``.

    Returns:
        DataFrame with columns ``segment, model, count, share, stichtag``.
    """
    if path is None:
        path = MODELLREIHEN_PATH
    counter = CoercionCounter("Modellreihen 2026 segment_model")
    df = _read_modellreihen(path)
    records = []
    unmapped_stock: dict[str, float] = {}
    mapped_stock = 0.0
    for _, row in df.iterrows():
        seg_key = segment_lookup_key(row["Segment"])
        canonical = KBA_SEGMENT_MAP.get(seg_key)
        count = _coerce_count(row["Anzahl"], counter)
        if canonical is None:
            label = str(row["Segment"]).strip()
            unmapped_stock[label] = unmapped_stock.get(label, 0.0) + count
            continue
        mapped_stock += count
        model = f"{str(row['Marke']).strip()} {str(row['Modellreihe']).strip()}"
        records.append({
            "segment": canonical,
            "model": model,
            "count": count,
            "stichtag": "2026-01-01",
        })
    counter.log()
    _report_unmapped_segments("extract_segment_model_2026", unmapped_stock,
                              mapped_stock)
    frame = pd.DataFrame(records)
    segment_totals = frame.groupby("segment")["count"].transform("sum")
    frame["share"] = frame["count"] / segment_totals
    return frame


# --------------------------------------------------------------------------- #
# Modellreihen -> kba_model_fuel.csv (per-model fuel weight)
# --------------------------------------------------------------------------- #
def extract_model_fuel(path=None) -> pd.DataFrame:
    """Modellreihen (2026): per-model fuel-type shares.

    Reads the KBA Modellreihen bestand CSV filtered to Berichtszeitpunkt
    ``01.01.2026`` and computes per-model powertrain shares:

    - ``petrol_share``: residual after all electrified/diesel counts,
      ``max(Anzahl - Diesel - Hybrid - BEV, 0) / Anzahl``.
      Note: the raw ``Hybrid`` column already **includes** ``Hybrid_Plugin``,
      so the full ``Hybrid`` is subtracted here (not the split non-plugin value).
    - ``diesel_share``: ``Diesel / Anzahl``.
    - ``hybrid_share``: ``max(Hybrid - Hybrid_Plugin, 0) / Anzahl``
      (non-plugin hybrid only).
    - ``phev_share``: ``Hybrid_Plugin / Anzahl``.
    - ``bev_share``: ``BEV / Anzahl``.

    Rows with ``Anzahl <= 0`` after coercion are skipped (logged).
    Rows whose ``Segment`` does not map via ``KBA_SEGMENT_MAP`` are skipped
    (logged); the skip count is included in the log message (no-silent-fallback).

    Task B6 consistency assertion: when the raw CSV carries a direct
    ``Hybrid_ohne_Plugin`` column (a KBA-reported non-plugin-hybrid count,
    distinct from the ``Hybrid``/``Hybrid_Plugin`` pair used above), this is
    compared against the COMPUTED ``hybrid_nonplugin = Hybrid - Hybrid_Plugin``
    for every row. A disagreement is logged as a WARNING with the row count --
    this validates our arithmetic against the source's own column but does
    NOT change the emitted value; ``hybrid_share`` always keeps the COMPUTED
    value (already covered by :mod:`tests.test_extract_kba_modellreihen`). The
    column is OPTIONAL: its absence (the case for every raw file seen so far)
    is silent, since it carries no information either way.

    The join key convention matches ``kba_segment_model.csv``:
    ``model = f"{Marke} {Modellreihe}"`` (uppercase as delivered by KBA).

    Args:
        path: Path to the raw Modellreihen CSV.  Defaults to
            ``MODELLREIHEN_PATH`` when ``None``.

    Returns:
        DataFrame with columns ``segment, model, stichtag, petrol_share,
        diesel_share, hybrid_share, phev_share, bev_share``.
    """
    if path is None:
        path = MODELLREIHEN_PATH
    counter = CoercionCounter("Modellreihen 2026 model_fuel")
    df = _read_modellreihen(path)
    has_hybrid_ohne_plugin = "Hybrid_ohne_Plugin" in df.columns
    records = []
    n_unmapped = 0
    n_zero_anzahl = 0
    n_hybrid_checked = 0
    n_hybrid_mismatch = 0
    for _, row in df.iterrows():
        canonical = KBA_SEGMENT_MAP.get(segment_lookup_key(row["Segment"]))
        if canonical is None:
            n_unmapped += 1
            continue
        model = f"{str(row['Marke']).strip()} {str(row['Modellreihe']).strip()}"
        anzahl = _coerce_count(row["Anzahl"], counter)
        if pd.isna(anzahl) or anzahl <= 0:
            n_zero_anzahl += 1
            continue
        diesel = np.nan_to_num(_coerce_count(row["Diesel"], counter))
        hybrid_all = np.nan_to_num(_coerce_count(row["Hybrid"], counter))
        hybrid_plugin = np.nan_to_num(_coerce_count(row["Hybrid_Plugin"], counter))
        bev = np.nan_to_num(_coerce_count(row["BEV"], counter))
        # Non-plugin hybrid = total hybrid minus plugin hybrid. This COMPUTED
        # value is the one emitted below (hybrid_share); it is already covered
        # by test_extract_model_fuel_hybrid_split and is never overwritten.
        hybrid_nonplugin = max(hybrid_all - hybrid_plugin, 0.0)
        if has_hybrid_ohne_plugin:
            hybrid_direct = _coerce_count(row["Hybrid_ohne_Plugin"], counter)
            if not pd.isna(hybrid_direct):
                n_hybrid_checked += 1
                if abs(hybrid_nonplugin - hybrid_direct) > 1e-6:
                    n_hybrid_mismatch += 1
        # Petrol residual: subtract full Hybrid (which already includes plugin).
        petrol = max(anzahl - diesel - hybrid_all - bev, 0.0)
        records.append({
            "segment": canonical,
            "model": model,
            "stichtag": "2026-01-01",
            "petrol_share": petrol / anzahl,
            "diesel_share": diesel / anzahl,
            "hybrid_share": hybrid_nonplugin / anzahl,
            "phev_share": hybrid_plugin / anzahl,
            "bev_share": bev / anzahl,
        })
    counter.log()
    total_skipped = n_unmapped + n_zero_anzahl
    logger.info(
        "[extract_model_fuel] models written=%d, skipped: unmapped segment=%d, "
        "zero/invalid Anzahl=%d (total skipped=%d)",
        len(records), n_unmapped, n_zero_anzahl, total_skipped,
    )
    if has_hybrid_ohne_plugin:
        if n_hybrid_mismatch > 0:
            logger.warning(
                "[extract_model_fuel] %d/%d row(s) where the computed non-plugin "
                "hybrid count (Hybrid - Hybrid_Plugin) disagrees with the source's "
                "own 'Hybrid_ohne_Plugin' column -- the COMPUTED value is kept "
                "(already tested); this is a consistency check against the source's "
                "own arithmetic, not a data correction.",
                n_hybrid_mismatch, n_hybrid_checked,
            )
        else:
            logger.info(
                "[extract_model_fuel] Hybrid_ohne_Plugin consistency check: all "
                "%d row(s) agree with the computed value (Hybrid - Hybrid_Plugin).",
                n_hybrid_checked,
            )
    return pd.DataFrame(records)


# --------------------------------------------------------------------------- #
# FZ 12.1 -> kba_segment_model.csv (DEPRECATED — superseded by 2026 Modellreihen)
# --------------------------------------------------------------------------- #
# Superseded by extract_segment_model_2026(); kept for reference/reversibility.


# --------------------------------------------------------------------------- #
# MiD segment x economic status, by region (Bundesland / Raumtyp)
# --------------------------------------------------------------------------- #
def extract_mid_segment_by_status(path: Path, segment_map: dict, region_name_map: dict,
                                  file_label: str) -> pd.DataFrame:
    """MiD 2023 segment x economic status, column-% per region.

    Each segment block: a "Basis: PKW" title, a region header row (col1 ==
    "Total", col2.. == region names), a "Basis gewichtet" row (weighted base),
    an "oekonomischer Status" marker, then the 5 status rows (sehr niedrig ..
    sehr hoch), each a column-% per region.  Long form output:
    ``region, segment, status, share_pct, base_weighted``.
    """
    # The MiD files contain a single sheet ("MiD Tabellen"); read it by index.
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet_name = workbook.sheetnames[0]
    workbook.close()
    rows = _read_sheet(path, sheet_name)
    counter = CoercionCounter(file_label)

    # Locate block starts (the segment title lines).
    block_starts = []
    for index, row in enumerate(rows):
        cell0 = _normalise_label(row[0]) if row[0] is not None else ""
        if cell0.lower().startswith("pkw-segmentierung nach kba"):
            # The segment name follows the last "- ".
            after = cell0.split("- ")[-1].strip().lower()
            canonical = segment_map.get(segment_lookup_key(after))
            block_starts.append((index, canonical, after))

    records = []
    for index, canonical, raw_name in block_starts:
        if canonical is None:
            logger.warning("[%s] unmapped MiD segment '%s' at row %d (skipped)",
                           file_label, raw_name, index)
            continue
        # Region header: the row that follows the "... | Total | <group> ..."
        # row.  In the "Total" row, col2 carries the group title (e.g.
        # "Bundesland"); the individual region names are in the *next* row's
        # col2 onward (col1 there is blank).
        header_row = None
        base_row = None
        status_rows = {}
        block_end = min(index + 26, len(rows))
        for offset in range(index, block_end):
            row = rows[offset]
            cell0 = _normalise_label(_cell(row, 0)).lower() if _cell(row, 0) is not None else ""
            cell1 = _normalise_label(_cell(row, 1)).lower() if _cell(row, 1) is not None else ""
            if header_row is None and cell1 == "total" and offset + 1 < block_end:
                header_row = rows[offset + 1]
            if cell0.startswith("basis gewichtet"):
                base_row = row
            if cell0 in MID_STATUS_MAP:
                status_rows[MID_STATUS_MAP[cell0]] = row
        if header_row is None or len(status_rows) != len(STATUS_LABELS):
            logger.warning("[%s] incomplete block for segment '%s' at row %d (skipped)",
                           file_label, canonical, index)
            continue
        # Region columns: every non-empty header cell from col2 onward.
        region_columns = []
        for col in range(2, len(header_row)):
            name = _cell(header_row, col)
            if name is not None and str(name).strip():
                wrapped = _normalise_label(name)
                clean = region_name_map.get(wrapped)
                if clean is None:
                    logger.warning("[%s] unmapped region header '%s' (kept verbatim)",
                                   file_label, wrapped)
                    clean = wrapped
                region_columns.append((col, clean))
        for col, region in region_columns:
            base_weighted = (_coerce_count(_cell(base_row, col), counter)
                             if base_row is not None else np.nan)
            for status in STATUS_LABELS:
                share_pct = _coerce_percent(_cell(status_rows[status], col), counter)
                records.append({
                    "region": region,
                    "segment": canonical,
                    "status": status,
                    "share_pct": share_pct,
                    "base_weighted": base_weighted,
                })
    counter.log()
    frame = pd.DataFrame(records)
    return frame


# --------------------------------------------------------------------------- #
# Regionalstatistik 46251-02 -> kba_kreis_fuel.csv
# --------------------------------------------------------------------------- #
def extract_kreis_fuel_46251(path: Path = FUEL_46251_PATH) -> pd.DataFrame:
    """Regionalstatistik 46251-02: per-Kreis Pkw by fuel (Stichtag 01.01.2025).

    Column order after the 3 id columns (stichtag, ags5, name): Insgesamt,
    Benzin, Diesel, Gas, Hybrid, darunter Plug-In-Hybrid, Elektro, sonstige.
    ``Hybrid`` INCLUDES the ``darunter PHEV`` subset, so non-plugin hybrid is
    ``Hybrid - PHEV``. ``Elektro`` == BEV. Dissolved Kreise carry ``-`` -> dropped.

    Task B3: the raw 46251-02 file covers EVERY German Kreis, not only the 8
    ZGB Kreise -- keeping every Kreis with valid counts lets cross-cordon
    in-commuters (who carry their real origin ``kreis_ags5``, see
    ``incommuters._incommuter_kreis_ags5``) draw their true home-Kreis fuel mix
    instead of falling back to the national one. The 8 ZGB Kreise keep the
    canonical ``ZGB_KREISE`` label (backward-compatible provenance, matches the
    other ZGB-only KBA tables); every other Kreis uses the file's own ``name``
    column, since it has no ``ZGB_KREISE`` entry.
    """
    counter = CoercionCounter("46251-02 kreis_fuel")
    raw = pd.read_csv(path, sep=";", skiprows=8, header=None, encoding="latin-1", dtype=str)
    cols = ["stichtag", "ags5", "name", "insg", "benzin", "diesel", "gas",
            "hybrid", "phev", "elektro", "sonstige"]
    raw = raw.iloc[:, :len(cols)]
    raw.columns = cols
    records = []
    n_zgb = 0
    n_non_zgb = 0
    for _, r in raw.iterrows():
        code = str(r["ags5"]).strip()
        insg = _coerce_count(r["insg"], counter)
        if pd.isna(insg) or insg <= 0:
            continue  # dissolved / suppressed Kreis
        petrol = _coerce_count(r["benzin"], counter)
        diesel = _coerce_count(r["diesel"], counter)
        gas = _coerce_count(r["gas"], counter)
        hybrid_all = _coerce_count(r["hybrid"], counter)
        phev = _coerce_count(r["phev"], counter)
        bev = _coerce_count(r["elektro"], counter)
        other = _coerce_count(r["sonstige"], counter)
        hybrid = max(np.nan_to_num(hybrid_all) - np.nan_to_num(phev), 0.0)
        if code in ZGB_KREISE:
            kreis_name = ZGB_KREISE[code]
            n_zgb += 1
        else:
            kreis_name = str(r["name"]).strip()
            n_non_zgb += 1
        records.append({
            "kreis_ags5": code, "kreis_name": kreis_name,
            "stichtag": "2025-01-01",
            "petrol": petrol, "diesel": diesel, "gas": gas, "bev": bev,
            "phev": phev, "hybrid": hybrid, "other": other,
        })
    counter.log()
    logger.info(
        "[extract_kreis_fuel_46251] %d Kreise kept (%d ZGB, %d non-ZGB); "
        "in-commuters from any of these Kreise now get their real home-Kreis "
        "fuel mix instead of the national fallback.",
        n_zgb + n_non_zgb, n_zgb, n_non_zgb,
    )
    frame = pd.DataFrame(records).sort_values("kreis_ags5").reset_index(drop=True)
    pts = ["petrol", "diesel", "gas", "bev", "phev", "hybrid", "other"]
    frame["total"] = frame[pts].sum(axis=1)
    for p in pts:
        frame[f"{p}_share"] = frame[p] / frame["total"]
    return frame


# --------------------------------------------------------------------------- #
# Regionalstatistik 46251-03 -> kba_kreis_euro.csv
# --------------------------------------------------------------------------- #
def extract_kreis_euro_46251(path: Path = EURO_46251_PATH) -> pd.DataFrame:
    """Regionalstatistik 46251-03: per-Kreis Pkw by Euro group (01.01.2025).

    Two rows per Kreis: ``insgesamt`` (all fuels) and ``Dieselangetriebener Pkw``.
    Columns after the 4 id cols (stichtag, ags5, name, teil): Insgesamt, Euro 1..5,
    Euro 6, darunter Euro-6d, darunter Euro-6d-temp, Sonstige.

    Task B4: the headline ``euro6`` column (the Destatis Euro-6 total) stays
    UNCHANGED, so every existing consumer (the per-Kreis Euro rake, ``total``,
    the ``euro*_share`` columns below) keeps working exactly as before. The two
    ``darunter`` columns -- previously read into the raw frame but discarded --
    are now ALSO emitted as additive count columns ``euro6d`` / ``euro6dtemp``
    (NOT part of ``euro_cols``/``total``, to avoid double counting the Euro-6
    total), plus the derived residual ``euro6ab = max(euro6 - euro6d -
    euro6dtemp, 0)`` (the pre-6d-temp Euro 6a/b/c share, not reported directly
    by Destatis). This mirrors the FZ 27.4 Euro-6 substage breakdown produced
    by :func:`extract_fuel_euro6_substage_nds`.

    Task B3: like :func:`extract_kreis_fuel_46251`, every Kreis covered by the
    raw 46251-03 file is kept (not only the 8 ZGB Kreise), so cross-cordon
    in-commuters get their real home-Kreis Euro-class mix. The 8 ZGB Kreise
    keep the canonical ``ZGB_KREISE`` label; every other Kreis uses the file's
    own ``name`` column.
    """
    counter = CoercionCounter("46251-03 kreis_euro")
    raw = pd.read_csv(path, sep=";", skiprows=8, header=None, encoding="latin-1", dtype=str)
    cols = ["stichtag", "ags5", "name", "teil", "insg",
            "e1", "e2", "e3", "e4", "e5", "e6", "e6d", "e6dtemp", "sonstige"]
    raw = raw.iloc[:, :len(cols)]
    raw.columns = cols
    records = []
    zgb_codes_kept: set = set()
    non_zgb_codes_kept: set = set()
    for _, r in raw.iterrows():
        code = str(r["ags5"]).strip()
        teil_raw = _normalise_label(r["teil"]).lower()
        if teil_raw.startswith("insgesamt"):
            teil = "all"
        elif teil_raw.startswith("dieselangetriebener"):
            teil = "diesel"
        else:
            continue
        insg = _coerce_count(r["insg"], counter)
        if pd.isna(insg) or insg <= 0:
            continue  # dissolved / suppressed Kreis (only ever occurs for non-ZGB Kreise here)
        euros = {k: _coerce_count(r[v], counter) for k, v in
                 (("euro1", "e1"), ("euro2", "e2"), ("euro3", "e3"),
                  ("euro4", "e4"), ("euro5", "e5"), ("euro6", "e6"),
                  ("other", "sonstige"))}
        # Task B4: Euro-6 substage counts.  "darunter Euro-6d" / "darunter
        # Euro-6d-temp" are SUBSETS of euro6 (raw values kept as-is, possibly
        # NaN for a KBA-suppressed cell, matching the sibling raw columns
        # above); the residual ("6a/b/c") is derived with a NaN-safe guard so a
        # suppressed darunter cell cannot turn the clamp negative.
        euro6d = _coerce_count(r["e6d"], counter)
        euro6dtemp = _coerce_count(r["e6dtemp"], counter)
        euro6ab = max(
            np.nan_to_num(euros["euro6"]) - np.nan_to_num(euro6d) - np.nan_to_num(euro6dtemp),
            0.0,
        )
        if code in ZGB_KREISE:
            kreis_name = ZGB_KREISE[code]
            zgb_codes_kept.add(code)
        else:
            kreis_name = str(r["name"]).strip()
            non_zgb_codes_kept.add(code)
        rec = {"kreis_ags5": code, "kreis_name": kreis_name,
               "stichtag": "2025-01-01", "teil": teil, **euros,
               "euro6d": euro6d, "euro6dtemp": euro6dtemp, "euro6ab": euro6ab}
        records.append(rec)
    counter.log()
    logger.info(
        "[extract_kreis_euro_46251] %d Kreise kept (%d ZGB, %d non-ZGB); "
        "in-commuters from any of these Kreise now get their real home-Kreis "
        "Euro-class mix instead of the national fallback.",
        len(zgb_codes_kept) + len(non_zgb_codes_kept),
        len(zgb_codes_kept), len(non_zgb_codes_kept),
    )
    frame = pd.DataFrame(records).sort_values(["kreis_ags5", "teil"]).reset_index(drop=True)
    # NOTE: euro_cols (and hence `total`/`*_share`) intentionally excludes
    # euro6d/euro6dtemp/euro6ab -- they are subsets/derivations of euro6, not
    # additional independent classes, so including them would double count.
    euro_cols = ["euro1", "euro2", "euro3", "euro4", "euro5", "euro6", "other"]
    frame["total"] = frame[euro_cols].sum(axis=1)
    for c in euro_cols:
        frame[f"{c}_share"] = frame[c] / frame["total"]
    return frame


# --------------------------------------------------------------------------- #
# KBA per-Gemeinde EV shares (2026.04) -> kba_gemeinde_ev.csv (ZGB only)
# --------------------------------------------------------------------------- #
def extract_gemeinde_ev(path: Path = GEMEINDE_EV_PATH) -> pd.DataFrame:
    """KBA per-Gemeinde EV shares (Stichtag 2026-04-01), BEV/PHEV/fuel-cell split.

    The KBA per-Gemeinde timeseries CSV (``kba_ev_gemeinde_timeseries_2023_2026.csv``)
    carries one row per Gemeinde per quarterly reporting period.  For ZGB Gemeinden
    the absolute Pkw counts are suppressed by KBA data-protection rules; only the
    ``*_Anteil`` (share) columns carry data -- the downstream tilt consumes shares,
    so this is loss-free for our use.

    This function:
    - Keeps only the latest reporting period (``Berichtszeitpunkt``, e.g. "2026.04").
    - Filters to Gemeinden whose AGS-8 prefix (first 5 digits) is one of the 8 ZGB
      Kreise (``ZGB_KREISE``).
    - Converts German decimal commas to dots and divides share columns by 100
      (percent -> fraction).
    - Logs the count of rows with any missing share values (no-silent-fallback rule);
      NaN shares are kept in the output so the consuming tilt can fall back to the
      Kreis-level value.
    - Applies ``normalize_gemeinde_name`` from
      ``braunschweig.synthesis.vehicles.fleet_sampling_de`` to produce the
      ``gemeinde_norm`` matching key -- the SAME function the population side
      uses, so the two vocabularies join.

    Args:
        path: Path to the raw KBA Gemeinde EV CSV (utf-8-sig, ``Berichtszeitpunkt``
              column carries the period string like ``"2026.04"``).

    Returns:
        DataFrame with columns: ``kreis_ags5, ags8, gemeinde, gemeinde_norm,
        stichtag, ev_share, bev_share, phev_share, fuelcell_share``.
        Sorted by (``kreis_ags5``, ``gemeinde``); index reset.
    """
    # Local import: the synthesis package is not guaranteed to be on sys.path
    # at script import time (only needed here, not by the other extractor functions).
    from braunschweig.synthesis.vehicles.fleet_sampling_de import normalize_gemeinde_name

    df = pd.read_csv(path, encoding="utf-8-sig", dtype=str)

    # Derive the 5-digit Kreis prefix from the 8-digit Gemeinde AGS.
    df["ags5"] = df["AGS"].str.strip().str[:5]

    # Keep only the latest reporting period.
    latest = sorted(df["Berichtszeitpunkt"].dropna().unique())[-1]  # e.g. "2026.04"
    stichtag = f"{latest[:4]}-{latest[5:7]}-01"

    sub = df[
        (df["Berichtszeitpunkt"] == latest) & (df["ags5"].isin(ZGB_KREISE))
    ].copy()

    def _frac(col: str) -> pd.Series:
        """Parse a percent share column (German comma) and divide by 100."""
        return (
            pd.to_numeric(
                sub[col].str.replace(",", ".", regex=False),
                errors="coerce",
            )
            / 100.0
        )

    out = pd.DataFrame(
        {
            "kreis_ags5": sub["ags5"].values,
            "ags8": sub["AGS"].str.strip().values,
            "gemeinde": sub["Gemeinde"].str.strip().values,
            "stichtag": stichtag,
            "ev_share": _frac("Pkw Elektro Anteil").values,
            "bev_share": _frac("Pkw_BEV_Anteil").values,
            "phev_share": _frac("Pkw Plug In Hybrid Anteil").values,
            "fuelcell_share": _frac("Pkw Brennstoffzelle Anteil").values,
        }
    )

    out["gemeinde_norm"] = out["gemeinde"].map(normalize_gemeinde_name)

    # Log missing-share rows (no-silent-fallback rule): if any share is NaN for a
    # ZGB row that is genuinely unexpected (not a KBA suppression), it is a signal
    # that the column mapping needs attention.  The consumer (EV tilt) will fall
    # back to the Kreis value for such rows.
    share_cols = ["ev_share", "bev_share", "phev_share", "fuelcell_share"]
    n_missing = int(out[share_cols].isna().any(axis=1).sum())
    n_total = len(out)
    if n_missing > 0:
        logger.warning(
            "[extract_gemeinde_ev] %d/%d ZGB Gemeinde rows have at least one missing "
            "share value (NaN kept; consuming tilt falls back to Kreis level).",
            n_missing, n_total,
        )
    else:
        logger.info(
            "[extract_gemeinde_ev] %d ZGB Gemeinde rows, all share columns populated "
            "(stichtag=%s).",
            n_total, stichtag,
        )

    return out.sort_values(["kreis_ags5", "gemeinde"]).reset_index(drop=True)


# --------------------------------------------------------------------------- #
# KBA 5 km EV-share grid (2026) -> kba_ev_grid.csv (ZGB bbox clip)
# --------------------------------------------------------------------------- #
def extract_ev_grid(path: Path = GRID_EV_PATH) -> pd.DataFrame:
    """KBA 5 km EV-share grid (2026), clipped to the ZGB bounding box.

    Reads the raw gpkg layer (EPSG:3857), clips cells to a generous ZGB
    bounding box (lon 10.0-11.7, lat 51.5-52.9, centroids reprojected to
    EPSG:4326), and returns one row per kept cell with:

    - ``cell_id``: raw ``id_5km`` string (e.g. "5kmN2695E4340").
    - ``stichtag``: "2026-04-01" (KBA reporting date).
    - ``ev_share``: ``elektro_an`` converted from percent to fraction; NaN where
      the value is missing or the cell is KBA-suppressed.
    - ``minx, miny, maxx, maxy``: cell geometry bounds in EPSG:3857.
    - ``suppressed``: True where ``ZS_Anteil_`` carries "-" (KBA low-count flag).

    geopandas is imported locally to avoid adding it to the module-level
    import path for callers that do not need gpkg support.

    Logs suppressed-cell and NaN-ev-share counts (no-silent-fallback rule).
    The consuming grid-tilt stage (a later task) falls back to the Gemeinde-level
    EV share for suppressed/NaN cells; this extractor never invents values.

    Args:
        path: Path to the raw KBA 5 km EV-share gpkg.  Defaults to
            ``GRID_EV_PATH``.

    Returns:
        DataFrame with columns
        ``cell_id, stichtag, ev_share, minx, miny, maxx, maxy, suppressed``.
        Sorted by ``cell_id``; index reset.
    """
    import geopandas as gpd

    gdf = gpd.read_file(str(path))  # geometry in EPSG:3857

    # Clip to a generous ZGB bounding box using WGS-84 centroids.
    cent = gdf.geometry.to_crs(4326).centroid
    keep = (cent.x >= 10.0) & (cent.x <= 11.7) & (cent.y >= 51.5) & (cent.y <= 52.9)
    g = gdf[keep].copy()

    bounds = g.geometry.bounds  # minx, miny, maxx, maxy in EPSG:3857

    suppressed_mask = g["ZS_Anteil_"].astype(str).str.strip() == "-"
    ev_share = pd.to_numeric(g["elektro_an"], errors="coerce") / 100.0

    out = pd.DataFrame({
        "cell_id": g["id_5km"].astype(str).values,
        "stichtag": "2026-04-01",
        "ev_share": ev_share.values,
        "minx": bounds["minx"].values,
        "miny": bounds["miny"].values,
        "maxx": bounds["maxx"].values,
        "maxy": bounds["maxy"].values,
        "suppressed": suppressed_mask.values,
    })

    n_total = len(out)
    n_suppressed = int(out["suppressed"].sum())
    n_nan = int(out["ev_share"].isna().sum())
    logger.info(
        "[extract_ev_grid] cells kept (ZGB clip): %d; suppressed (KBA '-'): %d; "
        "ev_share NaN: %d",
        n_total, n_suppressed, n_nan,
    )
    if n_nan > n_suppressed:
        # NaN cells beyond the suppressed set indicate a data-column issue.
        logger.warning(
            "[extract_ev_grid] %d NaN ev_share cells exceed suppressed count (%d); "
            "check that the 'elektro_an' column is populated in the source gpkg.",
            n_nan, n_suppressed,
        )

    return out.sort_values("cell_id").reset_index(drop=True)


# --------------------------------------------------------------------------- #
# KBA per-RegioStaR7 EV share timeseries -> kba_ev_regiostar7.csv
# (national LOGGING-ONLY cross-check; Task B6)
# --------------------------------------------------------------------------- #
def extract_ev_regiostar7(path: Path = EV_REGIOSTAR7_PATH) -> pd.DataFrame:
    """KBA per-RegioStaR7 EV share timeseries -> a national validation cross-check.

    Reads the KBA RegioStaR-7 EV timeseries CSV
    (``kba_ev_regiostar7_timeseries_2023_2026.csv``), one row per RegioStaR-7
    region per quarterly reporting period (``Berichtszeitpunkt``, e.g.
    ``"2026.04"``). This is a NATIONAL aggregate, not ZGB-specific, and is used
    ONLY by :func:`braunschweig.synthesis.vehicles.fleet_validation.crosscheck_ev_by_regiostar7`
    as a LOGGING-ONLY, order-of-magnitude cross-check -- it is never fed into
    the synthesis as an IPF control or asserted as a regional target (see the
    project's no-invented-reference-value rule).

    This function:

    - Keeps only the LATEST reporting period.
    - Maps ``Regiostar7 Nummer`` to the canonical ``rs7`` int code (71..77);
      the residual code ``99`` ("keine Zuordnung", unassigned) and any other
      unparseable code are DROPPED (logged, no-silent-fallback rule).
    - Converts German decimal commas in ``Pkw Elektro Anteil`` to dots and
      divides by 100 (percent -> fraction) to produce ``ev_share``.

    Args:
        path: Path to the raw KBA RegioStaR-7 EV timeseries CSV (utf-8-sig).
            Defaults to :data:`EV_REGIOSTAR7_PATH`.

    Returns:
        DataFrame with columns ``rs7, ev_share, stichtag``, one row per
        RegioStaR-7 code (71..77), sorted by ``rs7``.
    """
    df = pd.read_csv(path, encoding="utf-8-sig", dtype=str)

    # Keep only the latest reporting period (mirrors extract_gemeinde_ev).
    latest = sorted(df["Berichtszeitpunkt"].dropna().unique())[-1]  # e.g. "2026.04"
    stichtag = f"{latest[:4]}-{latest[5:7]}-01"

    sub = df[df["Berichtszeitpunkt"] == latest].copy()

    rs7_numeric = pd.to_numeric(sub["Regiostar7 Nummer"], errors="coerce")
    n_before = len(sub)
    keep = rs7_numeric.isin(RS7_CODES)
    n_dropped = int((~keep).sum())
    if n_dropped:
        logger.info(
            "[extract_ev_regiostar7] %d/%d row(s) dropped (RegioStaR7 code outside "
            "71..77, e.g. 99 'keine Zuordnung' or unparseable); %d kept.",
            n_dropped, n_before, n_before - n_dropped,
        )
    sub = sub[keep].copy()
    sub["rs7"] = rs7_numeric[keep].astype(int)

    ev_share = pd.to_numeric(
        sub["Pkw Elektro Anteil"].str.replace(",", ".", regex=False), errors="coerce",
    ) / 100.0
    n_nan = int(ev_share.isna().sum())
    if n_nan:
        logger.warning(
            "[extract_ev_regiostar7] %d/%d row(s) have an unparseable 'Pkw Elektro "
            "Anteil' value -> ev_share NaN.", n_nan, len(sub),
        )

    out = pd.DataFrame({
        "rs7": sub["rs7"].values,
        "ev_share": ev_share.values,
        "stichtag": stichtag,
    }).sort_values("rs7").reset_index(drop=True)

    logger.info(
        "[extract_ev_regiostar7] %d RegioStaR7 rows kept (stichtag=%s); national "
        "cross-check only, never an IPF control.",
        len(out), stichtag,
    )
    return out


# --------------------------------------------------------------------------- #
# Statista KBA ID 3438 -> kba_age_national.csv (VALIDATION control)
# --------------------------------------------------------------------------- #
def extract_age_national(path: Path = AGE_NATIONAL_PATH, year: int = 2026) -> pd.DataFrame:
    """KBA/Statista ID 3438: national Pkw age distribution (VALIDATION control).

    Sheet ``Daten``: the band-label row precedes the per-year rows (col1 is
    blank/misc, col2 = year value, cols 3.. = band percentages in the order
    listed in ``_AGE_NATIONAL_BANDS``).  Returns the requested year's 6 bands
    as a tidy DataFrame with columns ``year, stichtag, band, share_pct``.

    This is a national validation anchor (mean_age_years = 10.9 as stated in the
    Statista source); it is never used as an IPF dimension.

    Args:
        path: Path to the Statista xlsx (``Daten`` sheet).
        year: Reference year to extract (default 2026).

    Returns:
        DataFrame with 6 rows and columns ``year, stichtag, band, share_pct``.

    Raises:
        RuntimeError: If ``year`` is not found in the sheet.
    """
    rows = _read_sheet(path, "Daten")
    header: dict[int, str] | None = None
    for row in rows:
        labels = [(_normalise_label(c).lower() if c else "") for c in row]
        if "unter 2 jahre" in labels:
            header = {
                i: _AGE_NATIONAL_BANDS[labels[i]]
                for i in range(len(labels))
                if labels[i] in _AGE_NATIONAL_BANDS
            }
            continue
        if header is not None and row[1] is not None and str(row[1]).strip().isdigit() \
                and int(row[1]) == year:
            recs = [
                {"year": year, "stichtag": f"{year}-01-01",
                 "band": band, "share_pct": float(row[i])}
                for i, band in header.items()
            ]
            df = pd.DataFrame(recs)
            logger.info(
                "[age_national] year=%d, bands=%d, sum_share_pct=%.2f",
                year, len(df), df["share_pct"].sum(),
            )
            return df
    raise RuntimeError(f"age control: year {year} not found in {path}")


# --------------------------------------------------------------------------- #
# KBA wohnmobile holder-age distribution -> kba_wohnmobile_holder_age.csv
# (issue #315)
# --------------------------------------------------------------------------- #
def extract_wohnmobile_holder_age(path: Path = WOHNMOBILE_HOLDER_AGE_PATH) -> pd.DataFrame:
    """Wohnmobile stock by holder age (issue #315) -> kba_wohnmobile_holder_age.csv.

    The raw file is a hand transcription (no downloadable table exists behind
    the KBA infographic / PM 23/2025) and is COMMITTED; both source URLs, the
    Stichtag and the retrieval date live in its comment header. Validates the
    transcription -- counts including the ``not_attributed`` residual must sum
    exactly to the published total stock, so a typo fails loudly instead of
    silently re-weighting the tilt -- and adds ``share_of_attributed``: the age
    shares renormalised over the eight published natural-person classes
    (ADR-0093 ASSUMPTION: the unattributed residual carries the attributed age
    composition).
    """
    df = pd.read_csv(path, comment="#")
    required = ["age_class", "age_min_years", "age_max_years", "vehicles",
                "published_share_pct", "total_stock", "stichtag"]
    missing = set(required) - set(df.columns)
    if missing:
        raise RuntimeError(f"{path}: missing columns {sorted(missing)}.")
    labels = set(df["age_class"])
    expected = set(WOHNMOBILE_AGE_CLASS_LABELS) | {WOHNMOBILE_AGE_NOT_ATTRIBUTED}
    if labels != expected:
        raise RuntimeError(
            f"{path}: age_class labels {sorted(labels)} != expected "
            f"{sorted(expected)}."
        )
    totals = df["total_stock"].unique()
    if len(totals) != 1:
        raise RuntimeError(f"{path}: total_stock must be a single repeated value.")
    total_stock = int(totals[0])
    vehicles_sum = int(df["vehicles"].sum())
    if vehicles_sum != total_stock:
        raise RuntimeError(
            f"{path}: transcribed counts sum to {vehicles_sum}, published total "
            f"is {total_stock} -- transcription error; fix the raw CSV."
        )
    if len(df["stichtag"].unique()) != 1:
        raise RuntimeError(f"{path}: stichtag must be a single repeated value.")
    out = df.copy()
    att_mask = out["age_class"] != WOHNMOBILE_AGE_NOT_ATTRIBUTED
    att_total = float(out.loc[att_mask, "vehicles"].sum())
    out["share_of_attributed"] = np.where(
        att_mask, out["vehicles"] / att_total, np.nan)
    residual = total_stock - att_total
    logger.info(
        "[wohnmobile_holder_age] %d vehicles in the 8 natural-person classes "
        "(%.2f%% of the published stock %d); not_attributed residual %d "
        "(%.2f%%) renormalised away (ADR-0093 assumption).",
        int(att_total), 100.0 * att_total / total_stock, total_stock,
        int(residual), 100.0 * residual / total_stock,
    )
    return out[["age_class", "age_min_years", "age_max_years", "vehicles",
                "published_share_pct", "share_of_attributed", "total_stock",
                "stichtag"]]


# --------------------------------------------------------------------------- #
# Driver
# --------------------------------------------------------------------------- #
def _write(frame: pd.DataFrame, name: str) -> None:
    DERIVED_DIR.mkdir(parents=True, exist_ok=True)
    path = DERIVED_DIR / name
    frame.to_csv(path, index=False)
    logger.info("[write] %s (%d rows, %d cols)", path, len(frame), frame.shape[1])


def _write_with_header(frame: pd.DataFrame, name: str, header_line: str) -> None:
    """Write ``frame`` as CSV with a ``# ...`` comment line prepended.

    Mirrors ``_write`` but prefixes the file with ``header_line`` (which must
    start with ``#``) so provenance metadata (e.g. mean_age_years, source ID,
    stichtag) travels with the CSV.  The loaders already tolerate ``# ...``
    comment lines via ``pd.read_csv(..., comment="#")``.

    Args:
        frame: DataFrame to write.
        name: Output filename under ``DERIVED_DIR``.
        header_line: The comment line to prepend (e.g. ``"# key=val key2=val2"``).
    """
    DERIVED_DIR.mkdir(parents=True, exist_ok=True)
    path = DERIVED_DIR / name
    with open(path, "w", encoding="utf-8") as fh:
        fh.write(header_line + "\n")
        fh.write(frame.to_csv(index=False))
    logger.info("[write] %s (%d rows, %d cols, with header)", path, len(frame), frame.shape[1])


def main() -> None:
    for required in (FZ27_PATH, FZ12_PATH, MID_BUNDESLAND_PATH, MID_RAUMTYP_PATH,
                     FUEL_46251_PATH, EURO_46251_PATH, AGE_NATIONAL_PATH,
                     GEMEINDE_EV_PATH, MODELLREIHEN_PATH, GRID_EV_PATH,
                     WOHNMOBILE_HOLDER_AGE_PATH):
        if not required.exists():
            raise FileNotFoundError(
                f"Required raw KBA/MiD input missing: {required} "
                f"(raw xlsx are local-only; see {KBA_DIR / 'README.md'})."
            )

    df_segment_powertrain = extract_segment_powertrain()
    _write(df_segment_powertrain, "kba_segment_powertrain.csv")
    _write(extract_kreis_powertrain(), "kba_kreis_powertrain.csv")
    _write(extract_gemeinde_private_bev(), "kba_gemeinde_private_bev.csv")
    _write(extract_fuel_euro_nds(), "kba_fuel_euro_nds.csv")
    _write(extract_fuel_euro6_substage_nds(), "kba_fuel_euro6_substage_nds.csv")
    _write(extract_age_fuel(), "kba_age_fuel.csv")
    _write(extract_brand_powertrain(), "kba_brand_powertrain.csv")
    # kba_segment_model.csv is now produced from the 2026 Modellreihen source
    # (extract_segment_model_2026) instead of the legacy FZ 12.1 xlsx.
    df_segment_model = extract_segment_model_2026()
    # Fail before writing: a segment without model rows would silently emit cars
    # with an empty brand/model (issue #277).
    _check_segment_model_coverage(df_segment_model, df_segment_powertrain)
    _write(df_segment_model, "kba_segment_model.csv")
    _write(extract_model_fuel(), "kba_model_fuel.csv")
    _write(
        extract_mid_segment_by_status(MID_BUNDESLAND_PATH, MID_SEGMENT_MAP,
                                      MID_BUNDESLAND_NAME_MAP,
                                      "MiD bundesland segment_by_status"),
        "mid2023_segment_by_status_bundesland.csv",
    )
    _write(
        extract_mid_segment_by_status(MID_RAUMTYP_PATH, MID_SEGMENT_MAP,
                                      MID_RAUMTYP_NAME_MAP,
                                      "MiD raumtyp segment_by_status"),
        "mid2023_segment_by_status_raumtyp.csv",
    )
    _write(extract_kreis_fuel_46251(), "kba_kreis_fuel.csv")
    _write(extract_kreis_euro_46251(), "kba_kreis_euro.csv")
    _write_with_header(
        extract_age_national(),
        "kba_age_national.csv",
        "# mean_age_years=10.9 source=KBA/Statista ID3438 stichtag=2026-01-01",
    )
    _write(extract_gemeinde_ev(), "kba_gemeinde_ev.csv")
    _write(extract_ev_grid(), "kba_ev_grid.csv")
    _write(extract_wohnmobile_holder_age(), "kba_wohnmobile_holder_age.csv")

    # Task B6: the RegioStaR7 EV timeseries is a NEW, OPTIONAL raw input (a
    # national logging-only cross-check, never an IPF control) -- guard it
    # separately from the hard-required tuple above so main() still runs to
    # completion (and the other 15 derived CSVs still regenerate) before the
    # raw file has been supplied.
    if EV_REGIOSTAR7_PATH.exists():
        _write(extract_ev_regiostar7(), "kba_ev_regiostar7.csv")
    else:
        logger.info(
            "[main] %s absent -- skipping kba_ev_regiostar7.csv (Task B6 national "
            "EV cross-check input; optional, not required by any control). Supply "
            "the raw file and re-run to enable braunschweig.synthesis.vehicles."
            "fleet_validation.crosscheck_ev_by_regiostar7.",
            EV_REGIOSTAR7_PATH,
        )

    logger.info("[done] all KBA/MiD fleet reference CSVs written to %s", DERIVED_DIR)


if __name__ == "__main__":
    main()

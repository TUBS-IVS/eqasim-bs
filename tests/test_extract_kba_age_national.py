"""Tests for extract_age_national in scripts/extract_kba_fleet.py.

Uses a minimal openpyxl-built Daten-sheet fixture with 2 years of data;
asserts the 6 bands, the per-band values, the stichtag, and that shares sum
to 100 % (tolerance 0.2 ppt).
"""
import openpyxl
import pandas as pd
import pytest

import scripts.extract_kba_fleet as ex


def test_age_national_2026_bands(tmp_path):
    xlsx = tmp_path / "age.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daten"
    for _ in range(4):
        ws.append([None])
    ws.append([None, None, "unter 2 Jahre", "2 bis 4 Jahre", "5 bis 9 Jahre",
               "10 bis 14 Jahre", "15 bis 29 Jahre", "30 und mehr Jahre", None])
    for yr, vals in [(2025, [10.4, 13.8, 28.4, 20.6, 23.9, 2.9]),
                     (2026, [10.4, 13.3, 27.5, 21.0, 24.6, 3.1])]:
        ws.append([None, yr, *vals, "in %"])
    wb.save(xlsx)
    df = ex.extract_age_national(xlsx, year=2026)
    assert list(df["band"]) == ["under_2", "2_to_4", "5_to_9", "10_to_14", "15_to_29", "30_plus"]
    assert abs(df["share_pct"].sum() - 100.0) < 0.2
    assert (df["stichtag"] == "2026-01-01").all()


def test_age_national_correct_values(tmp_path):
    """Check that the 2026 share_pct values map exactly to the fixture values."""
    xlsx = tmp_path / "age.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daten"
    for _ in range(4):
        ws.append([None])
    ws.append([None, None, "unter 2 Jahre", "2 bis 4 Jahre", "5 bis 9 Jahre",
               "10 bis 14 Jahre", "15 bis 29 Jahre", "30 und mehr Jahre", None])
    ws.append([None, 2026, 10.4, 13.3, 27.5, 21.0, 24.6, 3.1, "in %"])
    wb.save(xlsx)
    df = ex.extract_age_national(xlsx, year=2026)
    expected = {"under_2": 10.4, "2_to_4": 13.3, "5_to_9": 27.5,
                "10_to_14": 21.0, "15_to_29": 24.6, "30_plus": 3.1}
    for _, row in df.iterrows():
        assert abs(row["share_pct"] - expected[row["band"]]) < 1e-6, (
            f"band {row['band']}: expected {expected[row['band']]}, got {row['share_pct']}"
        )


def test_age_national_year_not_found_raises(tmp_path):
    """RuntimeError if the requested year is absent from the sheet."""
    xlsx = tmp_path / "age.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daten"
    ws.append([None, None, "unter 2 Jahre", "2 bis 4 Jahre", "5 bis 9 Jahre",
               "10 bis 14 Jahre", "15 bis 29 Jahre", "30 und mehr Jahre", None])
    ws.append([None, 2025, 10.4, 13.8, 28.4, 20.6, 23.9, 2.9, "in %"])
    wb.save(xlsx)
    with pytest.raises(RuntimeError, match="year 2026 not found"):
        ex.extract_age_national(xlsx, year=2026)


def test_age_national_year_column(tmp_path):
    """year column in returned DataFrame equals the requested year."""
    xlsx = tmp_path / "age.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daten"
    ws.append([None, None, "unter 2 Jahre", "2 bis 4 Jahre", "5 bis 9 Jahre",
               "10 bis 14 Jahre", "15 bis 29 Jahre", "30 und mehr Jahre", None])
    ws.append([None, 2026, 10.4, 13.3, 27.5, 21.0, 24.6, 3.1, "in %"])
    wb.save(xlsx)
    df = ex.extract_age_national(xlsx, year=2026)
    assert (df["year"] == 2026).all()
